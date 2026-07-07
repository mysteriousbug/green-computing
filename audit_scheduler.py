"""
Audit Scheduling Assistant — v2
================================
Schedules H2 2026 audits given H1 commitments.

Inputs (in INPUT_DIR):
    leave_tracker.xlsx       - calendar format: row 1 month labels, row 2 day labels,
                               data rows = auditor + daily leave codes (A/A1/A2/B/T/V/S/O/P).
                               3 sheets (ICS, T&A, DM).
    skills.xlsx              - 3 sheets, cols: Auditor, Skills, Certifications
    planned_audits.xlsx      - SINGLE sheet, cols: Audit Number, TITLE,
                               Audit Primary Team, REPORT ISSUANCE PLANNED,
                               Analytics Used, Total Auditor Days, Primary Audit Days,
                               Reporting Quarter, Team Lead
    h1_allocations.xlsx      - calendar grid (same as scheduler grid output) showing
                               H1 2026 bookings. 3 sheets.
    interests.xlsx           - SINGLE sheet, cols: SL No, Audit Team, Audit Name,
                               Start Date, End Date, Preference 1, Preference 2, Preference 3
                               (each Preference cell holds comma-separated auditor names)

Outputs (in OUTPUT_DIR):
    assignments_grid.xlsx    - same week-grid format as h1_allocations.xlsx
    assignments_summary.xlsx - one row per audit
    warnings.xlsx            - warnings + auditor workload

Scheduling horizon: H2 2026 (1 Jul – 31 Dec 2026).
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------

INPUT_DIR = Path("./inputs")
OUTPUT_DIR = Path("./outputs")
CACHE_DIR = Path("./cache")
TEAMS = ["ICS", "T&A", "DM"]
SHEET_ALIASES = {
    "ICS": ["ICS", "ics", "Team ICS"],
    "T&A": ["T&A", "TA", "T and A", "Team T&A"],
    "DM": ["DM", "Team DM", "Data Management"],
}
SCHEDULE_YEAR = 2027
# Scheduling horizon — configurable (set via set_horizon() or the app quarter picker).
# Defaults to Q1 2027; the names H2_START/H2_END are retained for back-compat but now
# just hold whatever horizon is active.
HORIZON_START = date(2027, 1, 1)
HORIZON_END = date(2027, 3, 31)
H2_START = HORIZON_START     # back-compat aliases (many call sites use these names)
H2_END = HORIZON_END
TEAM_SIZE = 4
MIN_TEAM_MEMBERS = 0          # teams are optional; no enforced floor across teams
REPORTING_BUFFER_DAYS = 5
AUDIT_DAYS_PER_WEEK = 3                   # 1 working week ≈ 3 audit-days
MEMBER_FLEX_WEEKS = 2                     # legacy ± shift (kept for compatibility)
MEMBER_LEAD_DAYS = 90                     # members may start up to 3 months before report date
AVAILABILITY_THRESHOLD = 0.6

QUARTERS = {
    "Q1 2027": (date(2027, 1, 1), date(2027, 3, 31)),
    "Q2 2027": (date(2027, 4, 1), date(2027, 6, 30)),
    "Q3 2027": (date(2027, 7, 1), date(2027, 9, 30)),
    "Q4 2027": (date(2027, 10, 1), date(2027, 12, 31)),
    "H2 2026": (date(2026, 7, 1), date(2026, 12, 31)),
    "H1 2026": (date(2026, 1, 1), date(2026, 6, 30)),
}


def set_horizon(start: date, end: date) -> None:
    """Set the active scheduling window. Updates the back-compat aliases too."""
    global HORIZON_START, HORIZON_END, H2_START, H2_END, SCHEDULE_YEAR
    HORIZON_START, HORIZON_END = start, end
    H2_START, H2_END = start, end
    SCHEDULE_YEAR = start.year


# Leave codes: all of these mark unavailability.
LEAVE_CODES = {"A", "A1", "A2", "B", "T/V", "T", "V", "S", "O", "P"}

# Scoring (display-only composite)
W_SKILL = 0.35
W_PRIOR = 0.30
W_AVAIL = 0.20
W_INTEREST = 0.15

# Split
BASE_SPLIT = {"primary": 0.50, "second": 0.30, "third": 0.20}
TEAM_FLOOR = 0.15
# Analytics is an EXTERNAL team, not budgeted by us — it must NOT affect the split.
# (ANALYTICS_DM_BUMP retained at 0.0 for back-compat; do not raise it.)
ANALYTICS_DM_BUMP = 0.0
INFRA_TA_BUMP = 0.05
SECURITY_ICS_BUMP = 0.05

INFRA_KEYWORDS = {"aws", "azure", "gcp", "cloud", "network", "firewall", "segmentation",
                  "active directory", "ad ", "infrastructure", "datacenter", "datacentre",
                  "patching", "endpoint", "server"}
SECURITY_KEYWORDS = {"pam", "vault", "identity", "iam", "access", "privileged",
                     "red-team", "red team", "ransomware", "malware", "cyber", "ics",
                     "dlp", "encryption", "key management"}
DM_KEYWORDS = {"data", "analytics", "model", "ml ", "ai ", "report", "etl",
               "warehouse", "lineage", "quality"}

# ----------------------------------------------------------------------------
# DATA CLASSES
# ----------------------------------------------------------------------------

@dataclass
class Booking:
    audit_number: str
    audit_title: str
    start: date
    end: date
    source: str = "H2"


@dataclass
class Auditor:
    name: str
    role: str
    home_team: str
    skills: set[str] = field(default_factory=set)
    certifications: set[str] = field(default_factory=set)
    leaves: set[date] = field(default_factory=set)
    preferences: list[str] = field(default_factory=list)
    prior_audits: list[str] = field(default_factory=list)
    bookings: list[Booking] = field(default_factory=list)
    preference_satisfied: bool = False
    primary_domains: set[str] = field(default_factory=set)    # domains they lead
    secondary_domains: set[str] = field(default_factory=set)  # domains they back up
    skill_tokens: set[str] = field(default_factory=set)       # from skills.xlsx (tokenised)

    @property
    def is_sam(self) -> bool:
        return self.role.upper() == "SAM"

    @property
    def is_cosource(self) -> bool:
        r = self.role.upper().replace("-", "").replace(" ", "")
        return r in ("COSOURCE", "COSOURCED", "CS")

    @property
    def days_booked_h2(self) -> int:
        return sum(working_days_between(b.start, b.end)
                   for b in self.bookings if b.source == "H2")

    @property
    def assigned_h2_audits(self) -> list[str]:
        return [b.audit_number for b in self.bookings if b.source == "H2"]

    def is_busy_on(self, d: date) -> bool:
        if d in self.leaves:
            return True
        for b in self.bookings:
            if b.start <= d <= b.end:
                return True
        return False

    def has_overlap(self, start: date, end: date) -> bool:
        for b in self.bookings:
            if not (b.end < start or b.start > end):
                return True
        return False


@dataclass
class Audit:
    number: str
    title: str
    primary_team: str
    report_date: date
    analytics: bool
    total_days: int
    quarter: str
    primary_days: Optional[int] = None   # explicit primary-team allocation from input file
    suggested_lead: Optional[str] = None
    interest_start: Optional[date] = None
    interest_end: Optional[date] = None
    required_skills: set[str] = field(default_factory=set)
    required_certs: set[str] = field(default_factory=set)
    domain_tags: set[str] = field(default_factory=set)
    matched_domains: list[str] = field(default_factory=list)  # domains this audit maps to
    domain_override: Optional[str] = None    # explicit domain OR champion name from input
    tl_override: Optional[str] = None        # explicit TL name (from override column)
    match_confidence: float = 0.0            # 0..1; low = weak/guessed domain match
    weak_match: bool = False                 # True if TL came from a low-confidence match
    start_date: date = field(default=date(2027, 1, 1))
    end_date: date = field(default=date(2027, 1, 1))
    duration_calendar_days: int = 0
    team_days: dict[str, int] = field(default_factory=dict)
    team_headcount: dict[str, int] = field(default_factory=dict)
    assigned_tl: Optional[str] = None
    assigned_members: list[tuple[str, str, date, date]] = field(default_factory=list)


# ----------------------------------------------------------------------------
# UTILITIES
# ----------------------------------------------------------------------------

def working_days_between(start: date, end: date) -> int:
    days = 0
    d = start
    while d <= end:
        if d.weekday() < 5:
            days += 1
        d += timedelta(days=1)
    return days


def working_days_before(end: date, n: int) -> date:
    d = end
    while n > 0:
        d -= timedelta(days=1)
        if d.weekday() < 5:
            n -= 1
    return d


def monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _resolve_sheet(xl: pd.ExcelFile, team: str) -> str:
    for cand in SHEET_ALIASES[team]:
        if cand in xl.sheet_names:
            return cand
    for s in xl.sheet_names:
        if team.lower().replace(" ", "") in s.lower().replace(" ", ""):
            return s
    raise KeyError(f"No sheet for team {team} in {xl.io}")


def _parse_auditor_cell(raw) -> tuple[str, str]:
    s = str(raw).strip()
    if "/" in s:
        role, name = s.split("/", 1)
        return role.strip().upper(), name.strip()
    return "AM", s


def _parse_date_header(val) -> Optional[date]:
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.date()
    s = str(val).strip()
    for fmt in ("%A, %B %d, %Y", "%B %d, %Y", "%d %B %Y",
                "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).date()
    except (ValueError, TypeError):
        return None


# ----------------------------------------------------------------------------
# LEAVE TRACKER (calendar format)
# ----------------------------------------------------------------------------

def _parse_leave_sheet(df: pd.DataFrame) -> dict[str, tuple[str, set[date]]]:
    """Return {name: (role, set_of_leave_days)} from a calendar-format leave sheet."""
    out: dict[str, tuple[str, set[date]]] = {}
    auditor_row_idx = None
    for i in range(min(5, len(df))):
        row_vals = [str(v).strip().lower() for v in df.iloc[i].tolist()]
        if "auditor" in row_vals:
            auditor_row_idx = i
            break
    if auditor_row_idx is None:
        return out

    header = df.iloc[auditor_row_idx].tolist()
    date_columns: list[tuple[int, date]] = []
    for col_idx, val in enumerate(header):
        if col_idx == 0 or pd.isna(val):
            continue
        parsed = _parse_date_header(val)
        if parsed:
            date_columns.append((col_idx, parsed))
    if not date_columns:
        return out

    for r in range(auditor_row_idx + 1, len(df)):
        raw_name = df.iat[r, 0]
        if pd.isna(raw_name):
            continue
        role, name = _parse_auditor_cell(raw_name)
        if not name or name.lower() == "nan":
            continue
        days: set[date] = set()
        for col_idx, d in date_columns:
            cell = df.iat[r, col_idx]
            if pd.isna(cell):
                continue
            code = str(cell).strip().upper().replace(" ", "")
            if code in LEAVE_CODES:
                days.add(d)
        out[name] = (role, days)
    return out


# ----------------------------------------------------------------------------
# LOADERS
# ----------------------------------------------------------------------------

def load_auditors() -> dict[str, Auditor]:
    auditors: dict[str, Auditor] = {}

    # 1. Leave tracker DEFINES the roster (role + team + leaves).
    xl = pd.ExcelFile(INPUT_DIR / "leave_tracker.xlsx")
    for team in TEAMS:
        try:
            sheet = _resolve_sheet(xl, team)
        except KeyError:
            continue
        df = pd.read_excel(xl, sheet_name=sheet, header=None)
        for name, (role, days) in _parse_leave_sheet(df).items():
            auditors[name] = Auditor(name=name, role=role or "AM",
                                      home_team=team, leaves=days)

    # 2. Domain champions file attaches domain ownership.
    _load_domain_champions(INPUT_DIR / "domain_champions.xlsx", auditors)

    # 2b. Skills file attaches specialist skills (beyond-domain), if present.
    _load_skills(INPUT_DIR / "skills.xlsx", auditors)

    # 3. H1 allocations
    h1_path = INPUT_DIR / "h1_allocations.xlsx"
    if h1_path.exists():
        _load_h1_allocations(h1_path, auditors)

    # 4. Interests + window overrides
    _load_interests_and_overrides(INPUT_DIR / "interests.xlsx", auditors)

    return auditors


def _load_skills(path: Path, auditors: dict[str, Auditor]) -> None:
    """Optional skills.xlsx: single sheet, columns Auditor / Skills / Certifications.
    Auditor cell is 'role/name'. Skills are tokenised for matching against titles.
    Roster is NOT defined here — only attaches skills to existing auditors (and adds
    any missing person defensively)."""
    if not path.exists():
        return
    try:
        df = pd.read_excel(path)
    except Exception:
        return
    df.columns = [str(c).strip() for c in df.columns]

    def col(*cands):
        cl = {c.lower() for c in cands}
        for c in df.columns:
            if c.lower().strip() in cl:
                return c
        return None

    c_aud = col("auditor", "name", "auditor name")
    c_sk = col("skills", "skill")
    c_ce = col("certifications", "certification", "certs", "cert")
    if not c_aud:
        return

    for _, row in df.iterrows():
        raw = row.get(c_aud)
        if pd.isna(raw):
            continue
        role, name = _parse_auditor_cell(raw)
        if not name:
            continue
        aud = auditors.get(name)
        if aud is None:
            # person not in leave tracker — add defensively, no leave data
            aud = Auditor(name=name, role=role or "AM", home_team="ICS")
            auditors[name] = aud
        skills_text = ""
        if c_sk and pd.notna(row.get(c_sk)):
            skills_text = str(row[c_sk])
            aud.skills |= {s.strip().lower() for s in skills_text.split(",") if s.strip()}
        if c_ce and pd.notna(row.get(c_ce)):
            aud.certifications |= {s.strip().lower() for s in str(row[c_ce]).split(",") if s.strip()}
        # Tokenise skills for title matching (reuse the domain tokenizer)
        aud.skill_tokens |= _domain_tokens(skills_text)


# ----------------------------------------------------------------------------
# DOMAIN CHAMPIONS + DOMAIN↔TITLE MATCHING
# ----------------------------------------------------------------------------

# Registry of all known domains, populated when the champions file loads.
# Maps domain_name -> {"primary": [names], "secondary": [names], "team": str}
DOMAIN_REGISTRY: dict[str, dict] = {}

# Tokens too generic to drive a domain match.
_DOMAIN_STOPWORDS = {
    "and", "or", "of", "the", "a", "an", "for", "to", "in", "on", "as",
    "management", "services", "service", "audit", "review", "systems",
    "system", "control", "controls", "security", "inc", "including",
    "etc", "device", "devices", "and/or",
}


# Short tokens that are meaningful despite length (acronyms).
_KEEP_SHORT = {"ai", "ml", "pt", "ad", "os", "db", "vm", "dr", "ir"}
# Acronym → expansion tokens, so 'AI' matches 'artificial intelligence' domains.
_ACRONYM_EXPAND = {
    "ai": {"artificial", "intelligence"},
    "ml": {"machine", "learning"},
    "pqc": {"quantum", "cryptography", "cryptographic"},
    "genai": {"artificial", "intelligence", "generative"},
    "llm": {"language", "model"},
    "pam": {"privileged", "access"},
    "mfa": {"authentication", "factor"},
    "dlp": {"data", "leak", "loss"},
    "ctpi": {"third", "party", "intelligence", "client"},
    "ubia": {"business", "impact", "assessment"},
    "wrb": {"wrb"},
    "fmi": {"financial", "market", "infrastructure"},
    "caas": {"crypto", "cryptography"},
}


def _domain_tokens(text: str) -> set[str]:
    """Tokenise a domain name or audit title for overlap matching.
    Strips parenthetical example lists and generic stopwords, keeps meaningful
    acronyms, and expands common acronyms to their concept words."""
    # Keep parenthetical acronyms but drop long example lists: pull out short
    # all-caps tokens from parentheses before stripping them.
    paren_bits = re.findall(r"\(([^)]*)\)", text)
    text_main = re.sub(r"\(.*?\)", " ", text)
    raw = re.findall(r"[a-z0-9]+", text_main.lower())
    # add short acronyms found in parentheses (e.g. '(CTPI)', '(UBIA)')
    for bit in paren_bits:
        for tok in re.findall(r"[A-Za-z0-9]+", bit):
            if tok.isupper() and 2 <= len(tok) <= 6:
                raw.append(tok.lower())

    toks: set[str] = set()
    for t in raw:
        if t in _DOMAIN_STOPWORDS:
            continue
        if len(t) <= 2 and t not in _KEEP_SHORT:
            continue
        if t.endswith("s") and len(t) > 4:
            t = t[:-1]
        toks.add(t)
        if t in _ACRONYM_EXPAND:
            toks |= _ACRONYM_EXPAND[t]
    return toks


def _split_names(cell) -> list[str]:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return []
    s = str(cell).strip()
    if not s or s.lower() == "nan":
        return []
    return [n.strip() for n in s.split(",") if n.strip()]


def _load_domain_champions(path: Path, auditors: dict[str, Auditor]) -> None:
    """Single-sheet domain tracker:
       columns Team, Domain, Primary Product Champion, Secondary Product Champion.
       Primary champions are SAMs. Secondary cells may list multiple names."""
    DOMAIN_REGISTRY.clear()
    if not path.exists():
        return
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    def col(*cands):
        cands_l = {c.lower() for c in cands}
        for c in df.columns:
            if c.lower().strip() in cands_l:
                return c
        # loose contains-match
        for c in df.columns:
            for cand in cands:
                if cand.lower() in c.lower():
                    return c
        return None

    c_team = col("Team")
    c_domain = col("Domain")
    c_primary = col("Primary Product Champion", "Primary Champion", "Primary")
    c_secondary = col("Secondary Product Champion", "Secondary Champion", "Secondary")

    if not (c_domain and c_primary):
        return

    for _, row in df.iterrows():
        domain = row.get(c_domain)
        if pd.isna(domain):
            continue
        domain = str(domain).strip()
        if not domain:
            continue
        team = str(row.get(c_team, "")).strip() if c_team else ""
        primaries = _split_names(row.get(c_primary))
        secondaries = _split_names(row.get(c_secondary)) if c_secondary else []

        DOMAIN_REGISTRY[domain] = {
            "primary": primaries, "secondary": secondaries, "team": team,
        }
        for nm in primaries:
            if nm in auditors:
                auditors[nm].primary_domains.add(domain)
                # Primary champions are SAMs by definition.
                if not auditors[nm].is_sam:
                    auditors[nm].role = "SAM"
        for nm in secondaries:
            if nm in auditors:
                auditors[nm].secondary_domains.add(domain)


def match_domains_for_title(title: str, threshold: float = 0.18
                            ) -> list[tuple[str, float]]:
    """Return [(domain, score)] for domains whose tokens overlap the title,
    sorted best-first. Score = overlap / title-token-count (recall-biased)."""
    title_toks = _domain_tokens(title)
    if not title_toks:
        return []
    scored = []
    for domain, info in DOMAIN_REGISTRY.items():
        dtoks = _domain_tokens(domain)
        if not dtoks:
            continue
        overlap = title_toks & dtoks
        if not overlap:
            continue
        # recall over title tokens, lightly rewarded for covering domain tokens too
        score = len(overlap) / len(title_toks) + 0.25 * (len(overlap) / len(dtoks))
        scored.append((domain, score))
    scored.sort(key=lambda x: x[1], reverse=True)
    # Keep domains within range of the top score, above an absolute floor.
    if not scored:
        return []
    top = scored[0][1]
    return [(d, s) for d, s in scored if s >= max(threshold, top * 0.5)]


def _find_grid_headers(df: pd.DataFrame) -> tuple[Optional[int], Optional[int]]:
    for i in range(min(5, len(df))):
        row_vals = [str(v).strip().lower() for v in df.iloc[i].tolist()]
        if "auditor" in row_vals:
            return i, i
    return None, None


def _parse_week_label(val, year: int) -> Optional[date]:
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.date()
    s = str(val).strip()
    for fmt in ("%d/%b", "%d %b", "%d-%b", "%d/%B", "%d %B"):
        try:
            return datetime.strptime(s, fmt).date().replace(year=year)
        except ValueError:
            continue
    return _parse_date_header(val)


def _load_h1_allocations(path: Path, auditors: dict[str, Auditor]) -> None:
    xl = pd.ExcelFile(path)
    for team in TEAMS:
        try:
            sheet = _resolve_sheet(xl, team)
        except KeyError:
            continue
        df = pd.read_excel(xl, sheet_name=sheet, header=None)
        header_row, date_row = _find_grid_headers(df)
        if header_row is None:
            continue
        week_dates: list[tuple[int, date]] = []
        for col_idx in range(1, df.shape[1]):
            val = df.iat[date_row, col_idx]
            if pd.isna(val):
                continue
            wd = _parse_week_label(val, SCHEDULE_YEAR)
            if wd:
                week_dates.append((col_idx, wd))
        if not week_dates:
            continue

        for r in range(date_row + 1, len(df)):
            raw_name = df.iat[r, 0]
            if pd.isna(raw_name):
                continue
            _, name = _parse_auditor_cell(raw_name)
            if name not in auditors:
                continue
            current_title = None
            current_start = None
            current_end = None
            for col_idx, wstart in week_dates:
                cell = df.iat[r, col_idx]
                title = str(cell).strip() if pd.notna(cell) else ""
                title = re.sub(r"\(.*?\)", "", title).strip()
                if title == current_title and current_title:
                    current_end = wstart + timedelta(days=4)
                else:
                    if current_title:
                        auditors[name].bookings.append(Booking(
                            audit_number=f"H1-{current_title[:20]}",
                            audit_title=current_title,
                            start=current_start, end=current_end, source="H1",
                        ))
                        if current_title not in auditors[name].prior_audits:
                            auditors[name].prior_audits.append(current_title)
                    if title:
                        current_title = title
                        current_start = wstart
                        current_end = wstart + timedelta(days=4)
                    else:
                        current_title = None
            if current_title:
                auditors[name].bookings.append(Booking(
                    audit_number=f"H1-{current_title[:20]}",
                    audit_title=current_title,
                    start=current_start, end=current_end, source="H1",
                ))
                if current_title not in auditors[name].prior_audits:
                    auditors[name].prior_audits.append(current_title)


def _load_interests_and_overrides(path: Path, auditors: dict[str, Auditor]) -> dict:
    overrides: dict[str, tuple[Optional[date], Optional[date]]] = {}
    if not path.exists():
        _load_interests_and_overrides._overrides = overrides
        return overrides
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]

    def find_col(*candidates):
        cands = {x.lower() for x in candidates}
        for c in df.columns:
            if c.lower().strip() in cands:
                return c
        return None

    col_audit = find_col("Audit Name", "Audit", "Audit Number")
    col_start = find_col("Start Date", "Start")
    col_end = find_col("End Date", "End")
    pref_cols = [c for c in df.columns if c.lower().startswith("preference")]

    if col_audit is None:
        _load_interests_and_overrides._overrides = overrides
        return overrides

    for _, row in df.iterrows():
        audit_label = row.get(col_audit)
        if pd.isna(audit_label):
            continue
        audit_key = str(audit_label).strip().lower()
        sd = _parse_date_header(row[col_start]) if col_start and pd.notna(row.get(col_start)) else None
        ed = _parse_date_header(row[col_end]) if col_end and pd.notna(row.get(col_end)) else None
        if sd or ed:
            overrides[audit_key] = (sd, ed)
        for pc in pref_cols:
            cell = row.get(pc)
            if pd.isna(cell):
                continue
            for raw in str(cell).split(","):
                _, name = _parse_auditor_cell(raw)
                name = name.strip()
                if name in auditors and str(audit_label).strip() not in auditors[name].preferences:
                    auditors[name].preferences.append(str(audit_label).strip())

    _load_interests_and_overrides._overrides = overrides
    return overrides


def load_audits() -> list[Audit]:
    audits: list[Audit] = []
    xl = pd.ExcelFile(INPUT_DIR / "planned_audits.xlsx")
    df = pd.read_excel(xl, sheet_name=xl.sheet_names[0])
    df.columns = [str(c).strip() for c in df.columns]
    overrides = getattr(_load_interests_and_overrides, "_overrides", {})

    def find_col(*cands):
        cl = {c.lower() for c in cands}
        for c in df.columns:
            if c.lower().replace("-", " ").strip() in cl:
                return c
        return None

    col_primary_days = find_col("primary audit days", "primary auditor days",
                                 "primary team days", "primary days")
    col_domain = find_col("domain", "domain override", "override domain")
    col_report = find_col("report issuance planned", "report date",
                           "report issuance", "report")
    col_total = find_col("total auditor days", "total days", "auditor days")

    for _, row in df.iterrows():
        if pd.isna(row.get("Audit Number")):
            continue

        # Report date — tolerant. Missing -> None (validator flags it).
        report_date = None
        rd = row.get(col_report) if col_report else None
        if pd.notna(rd):
            try:
                report_date = (rd.date() if isinstance(rd, (datetime, pd.Timestamp))
                               else pd.to_datetime(rd, dayfirst=True).date())
            except (ValueError, TypeError):
                report_date = None

        analytics = str(row.get("Analytics Used", "")).strip().lower() in ("yes", "y", "true")
        tl = row.get("Team Lead")
        tl = None if pd.isna(tl) else str(tl).strip()

        # Total auditor days — tolerant. Non-numeric or missing -> 0 (validator flags).
        total_days = 0
        if col_total is not None:
            tv = row.get(col_total)
            if pd.notna(tv):
                try:
                    total_days = int(round(float(tv)))
                except (ValueError, TypeError):
                    total_days = 0

        primary_days = None
        if col_primary_days is not None:
            val = row.get(col_primary_days)
            if pd.notna(val):
                try:
                    primary_days = int(round(float(val)))
                except (ValueError, TypeError):
                    primary_days = None

        domain_override = None
        if col_domain is not None:
            dv = row.get(col_domain)
            if pd.notna(dv) and str(dv).strip():
                domain_override = str(dv).strip()

        audit = Audit(
            number=str(row["Audit Number"]).strip(),
            title=str(row.get("TITLE", "")).strip(),
            primary_team=str(row.get("Audit Primary Team", "")).strip(),
            report_date=report_date if report_date else date(SCHEDULE_YEAR, 1, 1),
            analytics=analytics,
            total_days=total_days,
            quarter=str(row.get("Reporting Quarter", "")).strip(),
            primary_days=primary_days,
            suggested_lead=tl,
            domain_override=domain_override,
        )
        # Flag missing report date via a sentinel the validator reads
        if report_date is None:
            audit.match_confidence = -1.0  # reused sentinel; validator special-cases
        for key in (audit.title.lower(), audit.number.lower()):
            if key in overrides:
                sd, ed = overrides[key]
                audit.interest_start = sd
                audit.interest_end = ed
                break
        audits.append(audit)

    # Filter to the active horizon (report date within, plus a small tail for
    # audits reporting just after the window that still need scheduling).
    return [a for a in audits
            if HORIZON_START <= a.report_date <= HORIZON_END + timedelta(days=60)]


# ----------------------------------------------------------------------------
# DURATION + SPLIT
# ----------------------------------------------------------------------------

def compute_duration(audit: Audit) -> None:
    if audit.interest_start and audit.interest_end:
        audit.start_date = audit.interest_start
        audit.end_date = audit.interest_end
    else:
        per_person_audit_days = max(1, round(audit.total_days / TEAM_SIZE))
        calendar_days = max(7, round(per_person_audit_days * 7 / AUDIT_DAYS_PER_WEEK))
        audit.end_date = working_days_before(audit.report_date, REPORTING_BUFFER_DAYS)
        audit.start_date = audit.end_date - timedelta(days=calendar_days - 1)
    # Clamp to H2 boundary: H2 audits cannot start before H2_START.
    if audit.start_date < HORIZON_START:
        audit.start_date = HORIZON_START
    audit.duration_calendar_days = (audit.end_date - audit.start_date).days + 1


def is_audit_compressed(audit: Audit) -> bool:
    """True if the H2 window is too short for the work required."""
    per_person_audit_days = max(1, round(audit.total_days / TEAM_SIZE))
    needed = round(per_person_audit_days * 7 / AUDIT_DAYS_PER_WEEK)
    return audit.duration_calendar_days < needed * 0.6


def compute_split(audit: Audit) -> None:
    title_l = audit.title.lower()
    primary = audit.primary_team
    others = [t for t in TEAMS if t != primary]
    share = {primary: BASE_SPLIT["primary"],
             others[0]: BASE_SPLIT["second"],
             others[1]: BASE_SPLIT["third"]}

    if audit.analytics and "DM" in share:
        share["DM"] += ANALYTICS_DM_BUMP
        donor = min((t for t in share if t != "DM"), key=lambda t: share[t])
        share[donor] -= ANALYTICS_DM_BUMP
    if any(kw in title_l for kw in INFRA_KEYWORDS) and "T&A" in share:
        donor = max((t for t in share if t != "T&A"), key=lambda t: share[t])
        share["T&A"] += INFRA_TA_BUMP
        share[donor] -= INFRA_TA_BUMP
    if any(kw in title_l for kw in SECURITY_KEYWORDS) and "ICS" in share:
        donor = max((t for t in share if t != "ICS"), key=lambda t: share[t])
        share["ICS"] += SECURITY_ICS_BUMP
        share[donor] -= SECURITY_ICS_BUMP

    for t in TEAMS:
        if share[t] < TEAM_FLOOR:
            deficit = TEAM_FLOOR - share[t]
            share[t] = TEAM_FLOOR
            donor = max(share, key=share.get)
            share[donor] -= deficit

    if audit.primary_days is not None:
        # Pin the primary team's allocation; distribute the remainder across the
        # other two teams in proportion to their heuristic shares.
        primary_days = max(0, min(audit.primary_days, audit.total_days))
        remainder = audit.total_days - primary_days
        other_shares = {t: share[t] for t in others}
        denom = sum(other_shares.values()) or 1.0
        raw = {primary: float(primary_days)}
        for t in others:
            raw[t] = remainder * (other_shares[t] / denom)
    else:
        raw = {t: share[t] * audit.total_days for t in TEAMS}

    rounded = {t: int(round(v)) for t, v in raw.items()}
    # Preserve the pinned primary value exactly before reconciling drift.
    if audit.primary_days is not None:
        rounded[primary] = max(0, min(audit.primary_days, audit.total_days))
    drift = audit.total_days - sum(rounded.values())
    if drift != 0:
        # When primary is pinned, only adjust the non-primary teams.
        adjustable = others if audit.primary_days is not None else TEAMS
        order = sorted(adjustable, key=lambda t: raw[t] - int(raw[t]), reverse=(drift > 0))
        for t in order:
            if drift == 0:
                break
            rounded[t] += 1 if drift > 0 else -1
            drift += -1 if drift > 0 else 1
    audit.team_days = rounded

    per_person = max(1, round(audit.total_days / TEAM_SIZE))
    audit.team_headcount = {
        t: max(1, round(rounded[t] / per_person)) if rounded[t] > 0 else 0
        for t in TEAMS
    }
    total = sum(audit.team_headcount.values())
    while total > TEAM_SIZE:
        donor = max((t for t in TEAMS if audit.team_headcount[t] > 1),
                    key=lambda t: audit.team_headcount[t], default=None)
        if donor is None:
            break
        audit.team_headcount[donor] -= 1
        total -= 1
    while total < MIN_TEAM_MEMBERS:
        recv = max(TEAMS, key=lambda t: audit.team_days[t])
        audit.team_headcount[recv] += 1
        total += 1


# ----------------------------------------------------------------------------
# SKILL INFERENCE
# ----------------------------------------------------------------------------

SKILL_CACHE_FILE = CACHE_DIR / "skill_inference.json"


def infer_skills_for_audits(audits: list[Audit]) -> None:
    """Map each audit title to domains. Honours an explicit Domain override:
       - If the override names a known domain, use it (confidence 1.0).
       - If it names a person (champion), stash as tl_override (confidence 1.0).
       Otherwise fall back to word-overlap matching and record confidence."""
    # Build a lowercase champion-name lookup for override resolution
    champion_names = set()
    for info in DOMAIN_REGISTRY.values():
        champion_names.update(n.lower() for n in info.get("primary", []))
        champion_names.update(n.lower() for n in info.get("secondary", []))

    for a in audits:
        a.tl_override = None
        if a.domain_override:
            ov = a.domain_override.strip()
            ov_l = ov.lower()
            # (1) exact/contains domain name?
            matched_dom = None
            for dom in DOMAIN_REGISTRY:
                if ov_l == dom.lower() or ov_l in dom.lower() or dom.lower() in ov_l:
                    matched_dom = dom
                    break
            if matched_dom:
                a.matched_domains = [matched_dom]
                a.match_confidence = 1.0
                a.weak_match = False
                a.domain_tags = {matched_dom}
                continue
            # (2) champion name?
            if ov_l in champion_names:
                a.tl_override = ov
                a.match_confidence = 1.0
                a.weak_match = False
                # also record the domains that person champions, for member fit
                doms = [d for d, info in DOMAIN_REGISTRY.items()
                        if ov_l in [n.lower() for n in info.get("primary", [])]
                        or ov_l in [n.lower() for n in info.get("secondary", [])]]
                a.matched_domains = doms
                a.domain_tags = set(doms)
                continue
            # (3) override text didn't resolve — treat as free-text domain hint
            matches = match_domains_for_title(ov)
            if matches:
                a.matched_domains = [d for d, _ in matches]
                a.match_confidence = matches[0][1] if matches else 0.0
                a.weak_match = a.match_confidence < 0.35
                a.domain_tags = set(a.matched_domains)
                continue

        # No override — word-overlap match on the title
        matches = match_domains_for_title(a.title)
        a.matched_domains = [d for d, _ in matches]
        a.match_confidence = matches[0][1] if matches else 0.0
        a.weak_match = (not matches) or a.match_confidence < 0.35
        a.domain_tags = set(a.matched_domains)


def skill_fit(auditor: Auditor, audit: Audit) -> float:
    """Word overlap between an auditor's specialist skills and the audit title.
    Returns 0..1. Used to lift skill-matched non-champions above generic filler."""
    if not auditor.skill_tokens:
        return 0.0
    title_toks = _domain_tokens(audit.title)
    # also consider matched-domain tokens so 'PQC' skill hits a 'Cryptography' audit
    for d in audit.matched_domains:
        title_toks |= _domain_tokens(d)
    if not title_toks:
        return 0.0
    overlap = auditor.skill_tokens & title_toks
    if not overlap:
        return 0.0
    return len(overlap) / len(auditor.skill_tokens | title_toks)


# ----------------------------------------------------------------------------
# AVAILABILITY + SCORING
# ----------------------------------------------------------------------------

def availability_in_window(auditor: Auditor, start: date, end: date) -> float:
    total = working_days_between(start, end)
    if total == 0:
        return 0.0
    busy = 0
    d = start
    while d <= end:
        if d.weekday() < 5 and auditor.is_busy_on(d):
            busy += 1
        d += timedelta(days=1)
    return max(0.0, (total - busy) / total)


def domain_fit(auditor: Auditor, audit: Audit) -> int:
    """Tiered domain fit:
       2 = primary champion of a matched domain
       1 = secondary champion of a matched domain
       0 = no domain fit"""
    md = set(audit.matched_domains)
    if auditor.primary_domains & md:
        return 2
    if auditor.secondary_domains & md:
        return 1
    return 0


def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def prior_match(auditor: Auditor, audit: Audit) -> float:
    if not auditor.prior_audits:
        return 0.0
    stop = {"audit", "of", "and", "the", "a"}
    cur_tokens = set(re.findall(r"\w+", audit.title.lower())) - stop
    best = 0.0
    for prior in auditor.prior_audits:
        prior_tokens = set(re.findall(r"\w+", prior.lower())) - stop
        best = max(best, jaccard(cur_tokens, prior_tokens))
    return best


def has_interest(auditor: Auditor, audit: Audit) -> bool:
    return (audit.number in auditor.preferences
            or audit.title in auditor.preferences)


def score(auditor: Auditor, audit: Audit,
          window: Optional[tuple[date, date]] = None) -> tuple[float, dict]:
    """Kept for display. 'skill' now reflects domain fit (0/0.5/1.0)."""
    skill = {2: 1.0, 1: 0.5, 0: 0.0}[domain_fit(auditor, audit)]
    prior = prior_match(auditor, audit)
    if window:
        avail = availability_in_window(auditor, *window)
    else:
        avail = availability_in_window(auditor, audit.start_date, audit.end_date)
    interest = 1.0 if has_interest(auditor, audit) else 0.0
    total = (W_SKILL * skill + W_PRIOR * prior
             + W_AVAIL * avail + W_INTEREST * interest)
    return total, {"skill": skill, "prior": prior, "avail": avail, "interest": interest}


# ----------------------------------------------------------------------------
# WINDOW PLACEMENT FOR MEMBERS
# ----------------------------------------------------------------------------

def _find_member_window(auditor: Auditor, audit: Audit, audit_days_share: int
                        ) -> Optional[tuple[date, date]]:
    desired_days = max(5, round(audit_days_share * 7 / AUDIT_DAYS_PER_WEEK))
    # Cap the slot to the audit's actual window length — useful for compressed audits.
    max_slot = audit.duration_calendar_days
    calendar_days = min(desired_days, max_slot)
    # Members may start up to 3 months before the report date (point 5),
    # but never before H2 begins.
    lead_start = audit.report_date - timedelta(days=MEMBER_LEAD_DAYS)
    earliest = max(HORIZON_START, lead_start)
    latest_start = audit.end_date - timedelta(days=calendar_days - 1)
    if latest_start < earliest:
        latest_start = earliest

    best_slot = None
    best_avail = -1.0
    cur = earliest
    while cur <= latest_start:
        slot_end = cur + timedelta(days=calendar_days - 1)
        if not auditor.has_overlap(cur, slot_end):
            avail = availability_in_window(auditor, cur, slot_end)
            if avail > best_avail:
                best_avail = avail
                best_slot = (cur, slot_end)
        cur += timedelta(days=7)

    if best_slot and best_avail >= AVAILABILITY_THRESHOLD:
        return best_slot
    return None


# ----------------------------------------------------------------------------
# SOLVER
# ----------------------------------------------------------------------------

def _member_sort_key(a: Auditor, audit: Audit):
    """Lexicographic ranking, best-first (so we negate for ascending sort):
       1. domain fit (primary > secondary > none)
       2. availability in the audit's broad window (fewer leaves = better)
       3. interest (nominated this audit)
       4. lighter existing H2 workload
    """
    fit = domain_fit(a, audit)
    # Skill-matched non-champions rank above generic filler (but below champions).
    sk = skill_fit(a, audit)
    skill_tier = 1 if (fit == 0 and sk > 0) else 0
    lead_start = max(HORIZON_START, audit.report_date - timedelta(days=MEMBER_LEAD_DAYS))
    avail = availability_in_window(a, lead_start, audit.end_date)
    interest = 1 if has_interest(a, audit) else 0
    workload = a.days_booked_h2
    # Order: domain fit, then skill tier, then availability, interest, workload.
    return (-fit, -skill_tier, -avail, -interest, workload)


def validate_inputs(auditors: dict[str, Auditor],
                    audits: list[Audit]) -> list[dict]:
    """Pre-flight data-quality checks. Returns a list of findings:
       {level: 'error'|'warning'|'info', audit: str|None, message: str}
    'error' = will produce wrong/blank output; 'warning' = review advised."""
    findings: list[dict] = []

    def add(level, msg, audit=None):
        findings.append({"level": level, "audit": audit, "message": msg})

    # Roster sanity
    if not auditors:
        add("error", "No auditors loaded — check the leave tracker file.")
    sams = [a for a in auditors.values() if a.is_sam]
    if not sams:
        add("error", "No SAMs in the roster — no audit can get a Team Lead.")

    # Champion names present in domain registry but missing from roster
    roster_lower = {n.lower() for n in auditors}
    for dom, info in DOMAIN_REGISTRY.items():
        for nm in info.get("primary", []) + info.get("secondary", []):
            if nm.lower() not in roster_lower:
                add("warning",
                    f"Champion '{nm}' (domain '{dom[:40]}') is not in the leave "
                    f"tracker — they can't be scheduled.")

    # Per-audit checks
    seen_titles = {}
    for a in audits:
        aid = a.number
        # Missing report date (sentinel set at load)
        if a.match_confidence == -1.0:
            add("error", f"Missing report date.", aid)
        # Total days
        if a.total_days <= 0:
            add("error", f"Total Auditor Days is missing or zero.", aid)
        # Primary team
        if a.primary_team not in TEAMS:
            add("error",
                f"Primary team '{a.primary_team}' is not one of {TEAMS}.", aid)
        # Primary days sanity
        if a.primary_days is not None and a.primary_days > a.total_days:
            add("warning",
                f"Primary Audit Days ({a.primary_days}) exceeds Total "
                f"({a.total_days}) — will be capped.", aid)
        # Domain match quality
        if a.domain_override:
            pass  # explicit override — trusted
        elif not a.matched_domains:
            add("warning",
                f"Title matched NO domain — TL will be a best-guess. "
                f"Add a Domain/Team Lead override.", aid)
        elif a.weak_match:
            add("warning",
                f"WEAK domain match (confidence {a.match_confidence:.2f}) — "
                f"review the assigned TL or add an override.", aid)
        # Duplicate detection
        key = (a.title.strip().lower(), a.primary_team, a.report_date)
        if key in seen_titles:
            add("warning",
                f"Looks like a duplicate of {seen_titles[key]} "
                f"(same title, team, date).", aid)
        else:
            seen_titles[key] = aid

    if not audits:
        add("error", "No audits fall within the selected horizon — "
                     "check report dates and the quarter selection.")

    return findings


def solve(auditors: dict[str, Auditor], audits: list[Audit]) -> list[str]:
    warnings: list[str] = []
    audits_sorted = sorted(audits, key=lambda a: a.report_date)

    for audit in audits_sorted:
        if is_audit_compressed(audit):
            warnings.append(
                f"{audit.number} ({audit.title}): COMPRESSED — window "
                f"{audit.duration_calendar_days} days is too short for "
                f"{audit.total_days} auditor-days. Consider splitting."
            )

        # ----- TL selection.
        # Priority: explicit TL/domain override → primary domain champion SAM →
        #           closest partial match. Weak matches are flagged.
        def tl_key(a: Auditor):
            fit = domain_fit(a, audit)
            avail = availability_in_window(a, audit.start_date, audit.end_date)
            return (-fit, -avail, a.days_booked_h2)

        tl_pool = [
            a for a in auditors.values()
            if a.is_sam
            and not a.has_overlap(audit.start_date, audit.end_date)
            and availability_in_window(a, audit.start_date, audit.end_date) >= AVAILABILITY_THRESHOLD
        ]
        chosen_tl = None

        # (1) explicit override: tl_override (resolved champion) or suggested_lead column
        override_name = audit.tl_override or audit.suggested_lead
        if override_name:
            # match against full roster, not just SAM pool (override is authoritative)
            for a in auditors.values():
                if a.name.lower() == override_name.lower():
                    chosen_tl = a
                    break
            if chosen_tl is None:
                warnings.append(
                    f"{audit.number} ({audit.title}): override TL '{override_name}' "
                    f"not found in roster — falling back to matching."
                )

        # (2) domain-champion SAM
        if chosen_tl is None:
            domain_sams = [a for a in tl_pool if domain_fit(a, audit) == 2]
            ranked = sorted(domain_sams or tl_pool, key=tl_key)
            chosen_tl = ranked[0] if ranked else None

        if chosen_tl is None:
            warnings.append(
                f"{audit.number} ({audit.title}): no available SAM "
                f"for {audit.start_date} → {audit.end_date}"
            )
            continue

        # Weak-match flag: low title→domain confidence AND no explicit override
        if audit.weak_match and not override_name:
            warnings.append(
                f"⚠️ {audit.number} ({audit.title}): WEAK domain match "
                f"(confidence {audit.match_confidence:.2f}) — TL {chosen_tl.name} is a "
                f"best-guess. Review, or set a Domain/Team Lead override."
            )
        elif domain_fit(chosen_tl, audit) < 2 and not override_name:
            warnings.append(
                f"{audit.number} ({audit.title}): TL {chosen_tl.name} is not a primary "
                f"domain champion for this title — closest available match used."
            )

        chosen_tl.bookings.append(Booking(
            audit_number=audit.number, audit_title=audit.title,
            start=audit.start_date, end=audit.end_date, source="H2",
        ))
        audit.assigned_tl = chosen_tl.name
        if has_interest(chosen_tl, audit):
            chosen_tl.preference_satisfied = True

        # ----- Co-source mandate (dormant unless co-sources exist in roster).
        # Any co-source in the audit's PRIMARY team who is available for the window
        # MUST be included. Team-locked: never cross-team. Counts toward budget.
        mandatory_cosources = [
            a for a in auditors.values()
            if a.is_cosource and a.home_team == audit.primary_team
            and a.name != chosen_tl.name
            and a.name not in {m[0] for m in audit.assigned_members}
        ]
        for cs in mandatory_cosources:
            # Availability check is defensive only (spec says they're never on leave).
            if cs.has_overlap(audit.start_date, audit.end_date):
                warnings.append(
                    f"⚠️ {audit.number} ({audit.title}): co-source {cs.name} is already "
                    f"booked elsewhere and cannot cover this audit (overlapping schedule)."
                )
                continue
            cs.bookings.append(Booking(
                audit_number=audit.number, audit_title=audit.title,
                start=audit.start_date, end=audit.end_date, source="H2",
            ))
            audit.assigned_members.append(
                (cs.name, cs.home_team, audit.start_date, audit.end_date))

    # ===== PASS 2: fill members, now that every TL is reserved =====
    for audit in audits_sorted:
        if audit.assigned_tl is None:
            continue  # no TL assigned in pass 1; skip member fill
        chosen_tl = auditors.get(audit.assigned_tl)
        # ----- Members. Teams are optional: fill each team's day budget where
        # people exist. Co-sources already added above count toward the budget.
        for team in TEAMS:
            need = audit.team_headcount.get(team, 0)
            if team == audit.primary_team:
                need -= 1   # TL occupies one primary-team slot
            # Subtract anyone already on this audit from this team (e.g. co-sources)
            already_this_team = sum(1 for m in audit.assigned_members if m[1] == team)
            need -= already_this_team
            if need <= 0:
                continue

            hc = max(1, audit.team_headcount.get(team, 1))
            audit_days_share = audit.team_days.get(team, 0) // hc
            already = {m[0] for m in audit.assigned_members}
            # Members must not be a TL of ANY audit (TLs are reserved for leading).
            reserved_tls = {a.assigned_tl for a in audits_sorted if a.assigned_tl}
            pool = [a for a in auditors.values()
                    if a.home_team == team and a.name != audit.assigned_tl
                    and a.name not in already
                    and a.name not in reserved_tls
                    and not a.is_cosource]   # co-sources handled by the mandate above
            pool.sort(key=lambda a: _member_sort_key(a, audit))

            picked = 0
            for cand in pool:
                if picked >= need:
                    break
                slot = _find_member_window(cand, audit, audit_days_share)
                if slot is None:
                    continue
                slot_start, slot_end = slot
                cand.bookings.append(Booking(
                    audit_number=audit.number, audit_title=audit.title,
                    start=slot_start, end=slot_end, source="H2",
                ))
                audit.assigned_members.append((cand.name, cand.home_team, slot_start, slot_end))
                if has_interest(cand, audit):
                    cand.preference_satisfied = True
                picked += 1

            if picked < need:
                warnings.append(
                    f"{audit.number} ({audit.title}): only filled {picked}/{need} from {team}"
                )

    unsatisfied = [a.name for a in auditors.values()
                   if a.preferences and not a.preference_satisfied]
    if unsatisfied:
        warnings.append(f"Preferred audit not assigned: {', '.join(unsatisfied)}")

    return warnings


# ----------------------------------------------------------------------------
# UPLOADED TRACKER PARSER
# ----------------------------------------------------------------------------

def _parse_tracker_date_header(val) -> Optional[date]:
    """Parse a week-start date header like '03 Aug 2026'."""
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.date()
    s = str(val).strip()
    for fmt in ("%d %b %Y", "%d %B %Y", "%d-%b-%Y", "%d/%m/%Y",
                "%Y-%m-%d", "%d %b", "%d-%b", "%d/%b"):
        try:
            d = datetime.strptime(s, fmt)
            if d.year < 2000:          # bare format without year: assume schedule year
                d = d.replace(year=SCHEDULE_YEAR)
            return d.date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).date()
    except (ValueError, TypeError):
        return None


# Cell values that mark leave rather than an audit assignment.
_TRACKER_LEAVE_TOKENS = LEAVE_CODES | {"leave", "annual leave", "block leave",
                                        "sick", "training", "public holiday",
                                        "volunteering", "other"}


def load_tracker_upload(
    path_or_buf,
    auditors: dict[str, Auditor],
    audits: list[Audit],
) -> tuple[dict[str, Auditor], list[Audit], list[str]]:
    """Parse an uploaded APS tracker and populate bookings on auditors/audits.

    Expected format (single sheet):
        Col A: Auditor Name
        Col B: Designation (AM / SAM / Co-Source)
        Col C: Team (ICS / T&A / DM)
        Col D+: week-start dates as headers ('03 Aug 2026'), cells = audit title(s)
                 or leave indicator; semicolon-separated for multiple audits.

    Returns:
        (auditors, audits, warnings)
    The auditors dict may include new names not in the leave tracker — they are added
    with role/team taken from the tracker columns.
    """
    warnings: list[str] = []
    df = pd.read_excel(path_or_buf, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    # Find the name / designation / team columns (tolerate name variants)
    def _fc(*cands) -> Optional[str]:
        cands_l = {c.lower() for c in cands}
        for c in df.columns:
            if c.lower().strip() in cands_l:
                return c
        return None

    col_name = _fc("auditor name", "auditor", "name")
    col_role = _fc("designation", "role", "title")
    col_team = _fc("team")

    if not col_name:
        warnings.append("Could not find 'Auditor Name' column in uploaded tracker.")
        return auditors, audits, warnings

    # Identify week-date columns
    week_cols: list[tuple[str, date]] = []
    skip = {col_name, col_role, col_team}
    for c in df.columns:
        if c in skip:
            continue
        d = _parse_tracker_date_header(c)
        if d:
            week_cols.append((c, d))

    if not week_cols:
        warnings.append("No date columns recognised in uploaded tracker. "
                         "Expected headers like '03 Aug 2026'.")
        return auditors, audits, warnings

    # Build a quick lookup: normalised title -> Audit object
    title_index: dict[str, Audit] = {
        a.title.strip().lower(): a for a in audits
    }

    # Clear all existing H2 bookings before loading from the tracker
    # (so the uploaded file is the sole source of truth for H2 state)
    for aud in auditors.values():
        aud.bookings = [b for b in aud.bookings if b.source != "H2"]
    for a in audits:
        a.assigned_tl = None
        a.assigned_members = []

    for _, row in df.iterrows():
        raw_name = row.get(col_name)
        if pd.isna(raw_name) or not str(raw_name).strip():
            continue
        name = str(raw_name).strip()

        # Role and team from columns; fall back to existing auditor record
        if col_role and pd.notna(row.get(col_role)):
            role = str(row[col_role]).strip()
        else:
            role = auditors[name].role if name in auditors else "AM"

        if col_team and pd.notna(row.get(col_team)):
            team = str(row[col_team]).strip()
        else:
            team = auditors[name].home_team if name in auditors else "ICS"

        # Upsert auditor
        if name not in auditors:
            auditors[name] = Auditor(name=name, role=role, home_team=team)
            warnings.append(f"New person in tracker (not in leave file): {name} — "
                             f"added as {role}/{team} with no leave data.")
        else:
            # Refresh role/team from tracker (tracker is authoritative)
            auditors[name].role = role
            auditors[name].home_team = team

        person = auditors[name]

        # Walk week columns and merge consecutive same-titled spans into bookings
        # Structure: {audit_title: (current_start, current_end)}
        open_spans: dict[str, tuple[date, date]] = {}

        def _flush(title: str, span: tuple[date, date]):
            s, e = span
            person.bookings.append(Booking(
                audit_number=f"TRACKER-{title[:20]}",
                audit_title=title, start=s, end=e + timedelta(days=4), source="H2",
            ))
            # Link to audit object if title matches
            a_obj = title_index.get(title.lower())
            if a_obj:
                if a_obj.assigned_tl is None and person.is_sam:
                    a_obj.assigned_tl = person.name
                elif person.name not in [m[0] for m in a_obj.assigned_members]:
                    a_obj.assigned_members.append(
                        (person.name, person.home_team, s, e + timedelta(days=4)))

        for col, wstart in week_cols:
            cell = row.get(col)
            if pd.isna(cell) or not str(cell).strip():
                # Gap — flush all open spans
                for t, span in list(open_spans.items()):
                    _flush(t, span)
                open_spans.clear()
                continue

            cell_str = str(cell).strip()
            # Check if this is a leave indicator
            if cell_str.lower() in _TRACKER_LEAVE_TOKENS:
                for t, span in list(open_spans.items()):
                    _flush(t, span)
                open_spans.clear()
                # Mark as leave if it's a standard code
                code = cell_str.upper().replace(" ", "")
                if code in LEAVE_CODES:
                    d = wstart
                    while d <= wstart + timedelta(days=4):
                        if d.weekday() < 5:
                            person.leaves.add(d)
                        d += timedelta(days=1)
                continue

            # One or more audit titles, semicolon-separated
            titles_in_cell = {t.strip() for t in cell_str.split(";") if t.strip()}

            # Titles no longer present in this cell → flush their spans
            for t in list(open_spans):
                if t not in titles_in_cell:
                    _flush(t, open_spans.pop(t))

            # Extend or start spans for titles in this cell
            for t in titles_in_cell:
                if t in open_spans:
                    open_spans[t] = (open_spans[t][0], wstart)   # extend
                else:
                    open_spans[t] = (wstart, wstart)             # new span

        # End of row — flush remaining open spans
        for t, span in open_spans.items():
            _flush(t, span)

    warnings_deduped = list(dict.fromkeys(warnings))
    return auditors, audits, warnings_deduped


# ----------------------------------------------------------------------------
# WRITERS
# ----------------------------------------------------------------------------

def write_grid(auditors: dict[str, Auditor], audits: list[Audit], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    first_monday = monday_of(date(SCHEDULE_YEAR, 1, 1))
    if first_monday.year < SCHEDULE_YEAR:
        first_monday += timedelta(days=7)
    weeks = [first_monday + timedelta(weeks=i) for i in range(52)]
    months = []
    for w in weeks:
        m = w.strftime("%b-%y")
        months.append(m if (not months or months[-1] != m) else "")

    for team in TEAMS:
        ws = wb.create_sheet(team)
        for i, m in enumerate(months, start=2):
            c = ws.cell(row=1, column=i, value=m)
            if m:
                c.font = Font(bold=True)
        ws.cell(row=2, column=1, value="Auditor").font = Font(bold=True)
        for i, w in enumerate(weeks, start=2):
            ws.cell(row=2, column=i, value=f"{w.day}/{w.strftime('%b')}").font = Font(bold=True)

        team_auditors = sorted(
            [a for a in auditors.values() if a.home_team == team],
            key=lambda x: (x.role != "SAM", x.role, x.name),
        )
        for r, aud in enumerate(team_auditors, start=3):
            ws.cell(row=r, column=1, value=f"{aud.role}/{aud.name}")
            for i, w in enumerate(weeks, start=2):
                week_end = w + timedelta(days=4)
                titles = [b.audit_title for b in aud.bookings
                          if b.start <= week_end and b.end >= w]
                if titles:
                    c = ws.cell(row=r, column=i, value="; ".join(titles))
                    is_h2 = any(b.source == "H2" and b.start <= week_end and b.end >= w
                                for b in aud.bookings)
                    c.fill = PatternFill("solid",
                                          start_color="C6E5B3" if is_h2 else "E8E8E8")
        ws.freeze_panes = "B3"
        ws.column_dimensions["A"].width = 28
        for i in range(2, len(weeks) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 14
    wb.save(path)


def write_summary(audits: list[Audit], auditors: dict[str, Auditor], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = ["Audit Number", "Title", "Primary Team", "Report Date",
               "Window", "Total Days", "Split (ICS/T&A/DM)",
               "Headcount (ICS/T&A/DM)", "Team Lead",
               "Members (name + window)", "Avg Skill Score", "Warnings"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="305496")
        c.alignment = Alignment(horizontal="center")
    for r, a in enumerate(audits, start=2):
        scores = []
        if a.assigned_tl and a.assigned_tl in auditors:
            scores.append(score(auditors[a.assigned_tl], a)[1]["skill"])
        for m_name, _, _, _ in a.assigned_members:
            if m_name in auditors:
                scores.append(score(auditors[m_name], a)[1]["skill"])
        avg = round(sum(scores)/len(scores), 2) if scores else 0.0
        warns = []
        if not a.assigned_tl: warns.append("no TL")
        if len(a.assigned_members) + (1 if a.assigned_tl else 0) < MIN_TEAM_MEMBERS:
            warns.append("understaffed")
        members_str = "; ".join(
            f"{n} [{s.strftime('%d %b')}–{e.strftime('%d %b')}]"
            for n, _, s, e in a.assigned_members)
        for col, val in enumerate([
            a.number, a.title, a.primary_team, a.report_date.isoformat(),
            f"{a.start_date.strftime('%d %b')} → {a.end_date.strftime('%d %b')}",
            a.total_days,
            f"{a.team_days.get('ICS',0)}/{a.team_days.get('T&A',0)}/{a.team_days.get('DM',0)}",
            f"{a.team_headcount.get('ICS',0)}/{a.team_headcount.get('T&A',0)}/{a.team_headcount.get('DM',0)}",
            a.assigned_tl or "—", members_str, avg, ", ".join(warns),
        ], start=1):
            ws.cell(row=r, column=col, value=val)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["J"].width = 60
    ws.freeze_panes = "A2"
    wb.save(path)


def write_warnings(warnings: list[str], auditors: dict[str, Auditor], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Warnings"
    ws.cell(row=1, column=1, value="Detail").font = Font(bold=True)
    for r, w in enumerate(warnings, start=2):
        ws.cell(row=r, column=1, value=w)
    ws.column_dimensions["A"].width = 100

    ws2 = wb.create_sheet("Workload (H2)")
    ws2.append(["Auditor", "Role", "Team", "H2 Audits", "H2 Days Booked", "Preference Met"])
    for a in sorted(auditors.values(), key=lambda x: (x.home_team, x.role, x.name)):
        ws2.append([
            a.name, a.role, a.home_team,
            len(a.assigned_h2_audits), a.days_booked_h2,
            "yes" if a.preference_satisfied else ("n/a" if not a.preferences else "no"),
        ])
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    wb.save(path)


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------

def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    print("Loading auditors...")
    auditors = load_auditors()
    print(f"  → {len(auditors)} auditors")
    print("Loading H2 audits...")
    audits = load_audits()
    print(f"  → {len(audits)} H2 audits")
    for a in audits:
        compute_duration(a); compute_split(a)
    infer_skills_for_audits(audits)
    warnings = solve(auditors, audits)
    print(f"  → {len(warnings)} warnings")
    write_grid(auditors, audits, OUTPUT_DIR / "assignments_grid.xlsx")
    write_summary(audits, auditors, OUTPUT_DIR / "assignments_summary.xlsx")
    write_warnings(warnings, auditors, OUTPUT_DIR / "warnings.xlsx")
    print("Done.")


if __name__ == "__main__":
    main()
